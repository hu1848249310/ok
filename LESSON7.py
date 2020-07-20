'''

''''''
接口自动化步骤：
1.Excel测试用例准备好，代码自动读取测试数据.  ---read_data
2.发送接口请求，得到响应信息---api_fun()
3.断言：实际结果VS预期结果---通过/不通过
4.写入通过/不通过---Excel  

注意：方便读取数据，Excel表格放到跟python文件同一级



Excel中的三大对象
工作簿workbook
表单sheet: cell=sheet.cell(row=2,column=1).value      #通过表单获取行号、列号---单元格
单元格cell
'''
from distutils.util import execute

import openpyxl
import requests
                               #导入库
#读取测试用例函数
def read_data(filename,sheetname):          #定义函数
    wb=openpyxl.load_workbook(filename)    #加载工作簿--文档名字
    sheet=wb[sheetname]     #获取表单
    max_row=sheet.max_row    #获取最大行数
    case_list=[]    #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1=dict(                               #打包成字典
        case_id=sheet.cell(row=i,column=1).value, #获取case_id(用例编号）
        url=sheet.cell(row=i,column=5).value,    #获取url
        data=sheet.cell(row=i,column=6).value,     #获取data
        expect=sheet.cell(row=i,column=7).value    #获取expect
        )
        case_list.append(dict1)                    #每循环一次，就把读取到的字典数据存放到这个list
    return case_list                                #返回测试用例列表

#执行接口参数
def api_fun(url,data):
    headers_login={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}    #请求头--字典
    result1=requests.post(url=url,json=data,headers=headers_login)     #接收post方法的结果
    response=result1.json()       #响应正文
    return response

#写入结果
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    sheet.cell(row=row,column=column).value=final_result                #写入结果
    wb.save('test_case_api.xlsx')                             #保存,关闭文档

#执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases=read_data(filename,sheetname)   #调用读取测试用例，获取所有测试用例，保存到一个变量中。
    for case in cases:
        case_id=case.get('case_id')                    #case.['case_id']
        url=case.get('url')
        data=eval(case.get('data'))                    #eval() 运行被字符串包裹的表达式---去掉字符串中的引号
        expect=eval(case.get('expect'))                #获取预期结果
        expect_msg=expect.get('msg')                   #获取预期结果中的msg
        real_result=api_fun(url=url,data=data)         #调用发送接口请求函数，返回结果用变量real_result接收
        real_msg=real_result.get('msg')                #获取实际结果中的msg
        print('预期结果中的msg：{}'.format(expect_msg))    #格式化输出
        print('实际结果中的msg：{}'.format(real_msg))
        if real_msg==expect_msg:
            print('第{}条测试用例执行通过！'.format(case_id) )                  #格式化输出
            final_result='Passed'                        #写入用例通过是的结果
        else:
            print('第{}条测试用例执行不通过！'.format(case_id))
            final_result='Failed'                         #写入用例不通过时的结果
        write_result(filename,sheetname,case_id+1,8,final_result)        #写入参数
        print('*'*40)

execute_fun('test_case_api.xlsx','register')









